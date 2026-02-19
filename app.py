# ── stdlib ─────────────────────────────────────────────────────────────────
import traceback
import datetime
from io import BytesIO
from pathlib import Path

# ── third-party ────────────────────────────────────────────────────────────
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Wedge


# ══════════════════════════════════════════════════════════════════════════
#  DQ ENGINE MODULES  (merged — replaces old split imports)
# ══════════════════════════════════════════════════════════════════════════
from modules.config          import AppConfig
from modules.case_management import page_case_management, init_case_management_state
from modules.ui_components   import UIComponents


# ══════════════════════════════════════════════════════════════════════════
#  PATCH — Fix broken HTML in UIComponents methods
#  The original render_lottie_upload produces raw ' style="color" text
#  visible in the UI. These replacements use safe, working HTML.
# ══════════════════════════════════════════════════════════════════════════
def _render_lottie_upload_fixed(caption: str = "Upload both files above to begin") -> None:
    st.markdown(
        """
        <div class="lottie-slot">
            <div class="lottie-frame lottie-upload-fallback"></div>
            <p class="lottie-caption">{caption}</p>
        </div>
        """.replace("{caption}", caption),
        unsafe_allow_html=True,
    )

def _render_arrow_down_fixed() -> None:
    st.markdown(
        '<div class="guidance-arrow-down">⬇</div>',
        unsafe_allow_html=True,
    )

def _render_upload_hint_fixed(kind: str = "dataset") -> None:
    if kind == "dataset":
        label = "📂 Master Dataset"
        tip   = "CSV, Excel (.xlsx/.xls/.xlsm), JSON, Parquet, ODS or XML"
    else:
        label = "📜 Rules / Rulebook"
        tip   = "CSV/Excel with column_name, rule, dimension, message — or a JSON rulebook"
    st.markdown(
        '<p style="font-size:0.82rem;color:#64748b;margin-bottom:0.3rem;">'
        + label + " &nbsp;·&nbsp; " + tip
        + "</p>",
        unsafe_allow_html=True,
    )

def _render_results_header_fixed(score: float) -> None:
    if score >= 80:
        cls, emoji, label = "dq-score-excellent", "🏆", "Excellent"
    elif score >= 60:
        cls, emoji, label = "dq-score-good",      "✅", "Good"
    elif score >= 40:
        cls, emoji, label = "dq-score-fair",      "⚠️", "Fair"
    else:
        cls, emoji, label = "dq-score-poor",      "❌", "Poor"
    st.markdown(
        '<div class="' + cls + '">'
        + '<h2 style="margin:0;">' + emoji + ' ' + label + ' — ' + f'{score:.1f}%' + '</h2>'
        + '</div>',
        unsafe_allow_html=True,
    )

# Patch UIComponents with fixed methods
UIComponents.render_lottie_upload  = _render_lottie_upload_fixed
UIComponents.render_arrow_down     = _render_arrow_down_fixed
UIComponents.render_upload_hint    = _render_upload_hint_fixed
UIComponents.render_results_header = _render_results_header_fixed

# data_quality_core  ←  dq_engine + rule_executor + rulebook_builder
from modules.data_quality_core import (
    RulebookBuilderService,
    RuleExecutorEngine,
    DataQualityEngine,
)

# reporting_core     ←  scoring_engine + report_generator
from modules.reporting_core import (
    ScoringService,
    ExcelReportGenerator,
)

# data_io_core       ←  file_loader + utils
from modules.data_io_core import (
    FileLoaderService,
    setup_directories,
    save_uploaded_file,
    clean_temp_directory,
)


# ══════════════════════════════════════════════════════════════════════════
#  DATA MATURITY MODULES
# ══════════════════════════════════════════════════════════════════════════
from DataMaturity.config import (
    UNIQU_PURPLE, UNIQU_MAGENTA, UNIQU_LAVENDER,
    UNIQU_LIGHT_BG, UNIQU_TEXT, UNIQU_GREY,
    RATING_LABELS, RATING_TO_SCORE,
    DEFAULT_MASTER_OBJECTS, MATURITY_DIMS, QUESTION_BANK,
)

from DataMaturity.helpers import (
    dq_score_to_maturity_level,
    init_maturity_state,
    build_question_df,
    sync_response_tables,
    autofill_dq_dimension,
    compute_all_scores,
    validate_responses,
    to_excel_bytes,
)

from DataMaturity.visualizations   import render_slide_png
from DataMaturity.report_generator import build_pdf_bytes

# ══════════════════════════════════════════════════════════════════════════
#  EXTERNAL CSS — assets/styles.css
# ══════════════════════════════════════════════════════════════════════════
def load_css():
    """Load external stylesheet from assets folder."""
    try:
        with open("assets/styles.css", encoding="utf-8") as f:
            st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
    except FileNotFoundError:
        st.warning("⚠️ styles.css not found in assets/ folder — place it at assets/styles.css")


# ══════════════════════════════════════════════════════════════════════════
#  FORCE FIX — DATA EDITOR DROPDOWN DARK THEME
# ══════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
div[data-baseweb="popover"],
div[data-baseweb="popover"] > div,
div[data-baseweb="menu"] {
    background: #ffffff !important;
    border: 1px solid #d9cef0 !important;
}
div[data-baseweb="popover"] ul,
div[data-baseweb="menu"] ul {
    background: #ffffff !important;
}
div[data-baseweb="popover"] li *,
div[data-baseweb="menu"] li *,
div[data-baseweb="popover"] [role="option"] *,
div[data-baseweb="menu"] [role="option"] * {
    color: #1a1a2e !important;
    -webkit-text-fill-color: #1a1a2e !important;
}
div[data-baseweb="popover"] [role="option"],
div[data-baseweb="menu"] [role="option"] {
    background: #1e293b !important;
    color: #1a1a2e !important;
    font-weight: 600 !important;
}
div[data-baseweb="popover"] [role="option"]:hover {
    background: #334155 !important;
}
div[data-baseweb="popover"] [aria-selected="true"] {
    background: rgba(124,58,237,0.35) !important;
    color: #e9d5ff !important;
}
div[data-baseweb="popover"] [data-highlighted] {
    background: rgba(96,165,250,0.25) !important;
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
#  GDG LIGHT THEME
#  Streamlit's Glide Data Grid reads --gdg-* vars from :root only,
#  so we inject them via st.markdown on every page that uses data_editor
# ══════════════════════════════════════════════════════════════════════════
_GDG_LIGHT_STYLE = """
<style>
/* ── Glide Data Grid: peaceful white/lavender theme ── */
:root,
[data-testid="stDataEditor"],
[data-testid="stDataEditor"] > div {
    --gdg-bg-cell:                 #ffffff !important;
    --gdg-bg-cell-medium:          #f7f4fc !important;
    --gdg-bg-header:               #ede8f7 !important;
    --gdg-bg-header-has-focus:     #e0d9f2 !important;
    --gdg-bg-header-hovered:       #d4cced !important;
    --gdg-border-color:            #e8e2f5 !important;
    --gdg-horizontal-border-color: #e8e2f5 !important;
    --gdg-accent-color:            #7c3aed !important;
    --gdg-accent-light:            rgba(124,58,237,0.10) !important;
    --gdg-text-dark:               #1a1028 !important;
    --gdg-text-medium:             #3b2f54 !important;
    --gdg-text-light:              #6b5f82 !important;
    --gdg-text-header:             #3b1f72 !important;
    --gdg-text-header-selected:    #1a0a40 !important;
    --gdg-cell-text-color:         #1a1028 !important;
    --gdg-header-font-style:       700 13px -apple-system, sans-serif !important;
    --gdg-base-font-style:         500 13px -apple-system, sans-serif !important;
}
[data-testid="stDataEditor"] canvas {
    background-color: #ffffff !important;
}
[data-testid="stDataEditor"] .dvn-scroller,
[data-testid="stDataEditor"] .dvn-scroll-inner,
[data-testid="stDataEditor"] > div,
[data-testid="stDataEditor"] > div > div {
    background: #ffffff !important;
}
[data-testid="stDataEditor"] ::-webkit-scrollbar        { width:7px; height:7px; }
[data-testid="stDataEditor"] ::-webkit-scrollbar-track  { background:#f3f0fa !important; }
[data-testid="stDataEditor"] ::-webkit-scrollbar-thumb  {
    background: #c4b5fd !important;
    border-radius: 4px;
}
[data-testid="stDataEditor"] ::-webkit-scrollbar-thumb:hover { background:#a78bfa !important; }
[data-testid="stDataEditor"] ::-webkit-scrollbar-corner      { background:#f3f0fa !important; }
[data-testid="stDataEditor"] input {
    background: #ffffff !important;
    color: #1a1028 !important;
    -webkit-text-fill-color: #1a1028 !important;
    border: 2px solid #7c3aed !important;
    border-radius: 5px !important;
    box-shadow: 0 0 0 3px rgba(124,58,237,0.12) !important;
}
.gdg-overlay-editor {
    background: #ffffff !important;
    border: 1.5px solid #d8d0ed !important;
    border-radius: 8px !important;
    box-shadow: 0 8px 28px rgba(91,45,144,0.15) !important;
    color: #1a1028 !important;
}
.gdg-overlay-editor select,
.gdg-overlay-editor select option {
    background: #ffffff !important;
    color: #1a1028 !important;
    -webkit-text-fill-color: #1a1028 !important;
}
</style>
"""


def inject_gdg_light():
    """Call once per page that renders st.data_editor."""
    st.markdown(_GDG_LIGHT_STYLE, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
#  SESSION STATE INITIALIZATION
# ══════════════════════════════════════════════════════════════════════════
def _init_state() -> None:
    # Navigation
    if "page" not in st.session_state:
        st.session_state["page"] = "home"

    # DQ results (populated after DQ run; consumed by Maturity)
    for key, default in {
        "dq_score":       None,
        "dq_dim_scores":  None,
        "dq_results_df":  None,
        "dq_object_name": "Customer",
        "dq_excel_path":  None,
    }.items():
        if key not in st.session_state:
            st.session_state[key] = default

    # Maturity state (uses keys defined in DataMaturity/helpers.py)
    init_maturity_state()

    # Policy Hub state
    if "policies" not in st.session_state:
        st.session_state["policies"] = []

    # Case Management state
    if "cases" not in st.session_state:
        st.session_state["cases"] = []


# ══════════════════════════════════════════════════════════════════════════
#  UTILITY FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════
def get_timestamp_filename(prefix: str, extension: str) -> str:
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.{extension}"


# ══════════════════════════════════════════════════════════════════════════
#  VISUALIZATION FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════
def _gauge_png(score: float) -> bytes:
    fig, ax = plt.subplots(figsize=(5, 3.2), dpi=150)
    fig.patch.set_facecolor('#ffffff')
    ax.set_xlim(0, 1); ax.set_ylim(0, 0.65); ax.axis("off")
    ax.add_patch(Wedge((0.5, 0.05), 0.40, 0, 180, width=0.12,
                       facecolor="#e5e7eb", edgecolor="white", lw=3))
    ang = score / 100 * 180
    col = "#10b981" if score >= 80 else ("#f59e0b" if score >= 60 else "#ef4444")
    ax.add_patch(Wedge((0.5, 0.05), 0.40, 0, ang, width=0.12,
                       facecolor=col, edgecolor="white", lw=3))
    ax.text(0.5, 0.32, f"{score:.1f}%", ha="center", va="center",
            fontsize=28, fontweight="bold", color="#6d28d9", family="sans-serif")
    ax.text(0.5, 0.18, "Overall DQ Score", ha="center", va="center",
            fontsize=11, color="#57534e", family="sans-serif", weight=600)
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", pad_inches=0.15, facecolor='#fafafa')
    plt.close(fig)
    return buf.getvalue()


def _dim_bar_png(dim_scores: dict) -> bytes | None:
    if not dim_scores:
        return None
    dims   = list(dim_scores.keys())
    scores = [dim_scores[d] for d in dims]
    cols   = ["#10b981" if s >= 80 else ("#f59e0b" if s >= 60 else "#ef4444") for s in scores]
    fig, ax = plt.subplots(figsize=(8, max(3, len(dims) * 0.8)), dpi=140)
    fig.patch.set_facecolor('#ffffff'); ax.set_facecolor('#f9f8fc')
    bars = ax.barh(dims, scores, color=cols, height=0.6, edgecolor="white", linewidth=2)
    ax.set_xlim(0, 112)
    ax.set_xlabel("DQ Score (%)", color="#1a1a2e", fontsize=11, weight=600, family="sans-serif")
    ax.tick_params(colors="#44403c", labelsize=10)
    ax.spines[["top", "right", "bottom"]].set_visible(False)
    ax.spines["left"].set_color("#d9cef0"); ax.spines["left"].set_linewidth(1.5)
    ax.axvline(80, color="#6d28d9", lw=1.5, ls="--", alpha=0.6, label="Excellent (80%)")
    ax.axvline(60, color="#7c3aed", lw=1.5, ls=":",  alpha=0.6, label="Good (60%)")
    ax.legend(fontsize=9, loc="lower right", frameon=True, fancybox=True, shadow=True)
    for bar, sc in zip(bars, scores):
        ax.text(bar.get_width() + 2, bar.get_y() + bar.get_height() / 2,
                f"{sc:.1f}%", va="center", fontsize=10,
                fontweight="bold", color="#1a1a2e", family="sans-serif")
    plt.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor='#ffffff')
    plt.close(fig)
    return buf.getvalue()


def _mat_bar_png(dim_vals: dict) -> bytes | None:
    if not dim_vals:
        return None
    dims   = list(dim_vals.keys())
    scores = [dim_vals[d] for d in dims]
    cols   = ["#5b2d90" if s >= 4 else "#7c4dbb" if s >= 3 else "#c4b0e0" for s in scores]
    fig, ax = plt.subplots(figsize=(10, max(3, len(dims) * 0.9)), dpi=140)
    fig.patch.set_facecolor('#f5f0fc'); ax.set_facecolor('#f9f8fc')
    bars = ax.barh(dims, scores, color=cols, height=0.6, edgecolor="white", linewidth=2)
    ax.set_xlim(0, 6.0)
    ax.set_xlabel("Maturity Score (1 = Adhoc  →  5 = Optimised)",
                  color="#3d1d63", fontsize=11, weight=600, family="sans-serif")
    ax.axvline(3.0, color="#38bdf8", lw=1.5, ls="--", alpha=0.7, label="Defined (3)")
    ax.axvline(4.0, color="#0284c7", lw=1.5, ls="--", alpha=0.7, label="Managed (4)")
    ax.legend(fontsize=9, loc="lower right", frameon=True, fancybox=True, shadow=True)
    ax.tick_params(colors="#0c4a6e", labelsize=10)
    ax.spines[["top", "right", "bottom"]].set_visible(False)
    ax.spines["left"].set_color("#c4b0e0"); ax.spines["left"].set_linewidth(2)
    for bar, sc in zip(bars, scores):
        ax.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height() / 2,
                f"{sc:.2f}", va="center", fontsize=11,
                fontweight="bold", color="#3d1d63", family="sans-serif")
    plt.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor='#f5f0fc')
    plt.close(fig)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════
#  COMBINED EXCEL (DQ + Maturity)
# ══════════════════════════════════════════════════════════════════════════
def _combined_excel(dq_score: float, dq_dim_scores: dict | None, mat_excel: bytes) -> bytes:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = load_workbook(BytesIO(mat_excel))
    header_fill = PatternFill(start_color="6d28d9", end_color="7c3aed", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    ws_dq = wb.create_sheet("DQ Score Summary", 0)
    ws_dq.append(["Metric", "Value"])
    for cell in ws_dq[1]:
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_dq.append(["Overall DQ Score (%)",  f"{dq_score:.1f}%"])
    ws_dq.append(["Mapped Maturity Level", dq_score_to_maturity_level(dq_score)])
    if dq_dim_scores:
        for dim, sc in dq_dim_scores.items():
            ws_dq.append([f"DQ – {dim}", f"{sc:.1f}%"])
    ws_dq.column_dimensions['A'].width = 30
    ws_dq.column_dimensions['B'].width = 20

    dq_df = st.session_state.get("dq_results_df")
    if dq_df is not None:
        display_cols = [c for c in dq_df.columns if not c.startswith("_")]
        ws_res = wb.create_sheet("DQ Results", 1)
        ws_res.append(display_cols)
        for cell in ws_res[1]:
            cell.fill = header_fill; cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for _, row in dq_df[display_cols].head(1000).iterrows():
            ws_res.append([str(v) if v is not None else "" for v in row.tolist()])

    out = BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()

# ══════════════════════════════════════════════════════════════════════════
#  UNIQUS TOP BAR  (shown on every page)
# ══════════════════════════════════════════════════════════════════════════
def render_uniqus_topbar(page_label: str = "Data Quality Intelligence Studio") -> None:
    st.markdown(f"""
    <div class="uniqus-topbar">
        <div class="uniqus-topbar-brand">
            <div class="uniqus-topbar-logo">U</div>
            <div>
                <div class="uniqus-topbar-title">Uniqus Consultech</div>
                <div class="uniqus-topbar-subtitle">{page_label}</div>
            </div>
        </div>
        <div class="uniqus-topbar-pill">Data Intelligence Studio</div>
    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
#  STATIC PBIX-STYLE DASHBOARD  (Data Quality – Executive Outlook)
#  Mirrors the Power BI report structure:
#  • 4 KPI cards (CDEs, Active Records, Rules, Avg DQ Score)
#  • 4 Gauge SVGs (Completeness, Standardization, Uniqueness, Validation)
#  • Stacked bar (Overall DQ per master data)
#  • Funnel chart (Rules by master data)
#  • Detail table (score by system/column)
# ══════════════════════════════════════════════════════════════════════════
def _svg_gauge(score: float, label: str, color: str = "#5b2d90") -> str:
    """Return an inline SVG semi-circular gauge."""
    pct   = max(0.0, min(100.0, score))
    # arc from 180° to 0°  (left to right)
    r     = 45
    cx, cy = 60, 55
    circ  = 3.14159 * r  # half circumference
    dash  = pct / 100 * circ
    gap   = circ - dash
    # colour by threshold
    if score >= 80:
        col = "#5b2d90"
    elif score >= 60:
        col = "#b10f74"
    elif score >= 40:
        col = "#d97706"
    else:
        col = "#dc2626"
    return f"""
    <div style="text-align:center;">
        <svg width="120" height="70" viewBox="0 0 120 70" xmlns="http://www.w3.org/2000/svg">
            <!-- background arc -->
            <path d="M 15 55 A {r} {r} 0 0 1 105 55"
                  fill="none" stroke="#e9e4f5" stroke-width="10"
                  stroke-linecap="round"/>
            <!-- foreground arc -->
            <path d="M 15 55 A {r} {r} 0 0 1 105 55"
                  fill="none" stroke="{col}" stroke-width="10"
                  stroke-linecap="round"
                  stroke-dasharray="{dash:.1f} {gap:.1f}"
                  stroke-dashoffset="0"/>
            <text x="60" y="50" text-anchor="middle"
                  font-size="15" font-weight="800" fill="{col}"
                  font-family="Inter, sans-serif">{score:.0f}%</text>
        </svg>
        <div style="font-size:0.72rem;font-weight:600;color:#7a7a9a;
                    text-transform:uppercase;letter-spacing:0.05em;margin-top:-6px;">{label}</div>
    </div>"""


def render_static_dq_dashboard(
    overall: float,
    dim_scores: dict,
    results_df,
    col_scores: dict,
) -> None:
    """
    Render a Power BI-style static executive dashboard matching the
    Smartworks Data Quality Assessment PBIX report layout.
    """
    # ── top bar ──────────────────────────────────────────────────────────
    render_uniqus_topbar("Data Quality Assessment — Executive Outlook")

    # ── derive quick stats ────────────────────────────────────────────────
    n_records  = len(results_df) if results_df is not None else 0
    n_rules    = len([c for c in results_df.columns if not c.startswith("_")]) if results_df is not None else 0
    clean_pct  = 0.0
    if results_df is not None and "Count of issues" in results_df.columns:
        clean_pct = round(len(results_df[results_df["Count of issues"] == 0]) / max(len(results_df), 1) * 100, 1)
    n_cols     = len(col_scores) if col_scores else 0
    avg_score  = overall

    badge_cls  = "excellent" if avg_score >= 80 else ("good" if avg_score >= 60 else ("fair" if avg_score >= 40 else "poor"))

    # ── KPI cards ─────────────────────────────────────────────────────────
    st.markdown("""
    <div class="dash-section-header">
        <div class="dash-section-dot"></div>
        <h3>Key Performance Indicators</h3>
        <div class="dash-section-accent"></div>
    </div>
    """, unsafe_allow_html=True)

    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.markdown(f"""
        <div class="exec-kpi-card purple">
            <span class="exec-kpi-icon">📊</span>
            <div class="exec-kpi-label">Number of CDEs</div>
            <div class="exec-kpi-value">{n_cols}</div>
            <div class="exec-kpi-delta">Critical Data Elements analyzed</div>
            <span class="exec-kpi-badge purple">CDEs</span>
        </div>""", unsafe_allow_html=True)
    with k2:
        st.markdown(f"""
        <div class="exec-kpi-card teal">
            <span class="exec-kpi-icon">🗂️</span>
            <div class="exec-kpi-label">Number of Active Records</div>
            <div class="exec-kpi-value teal">{n_records:,}</div>
            <div class="exec-kpi-delta">Records processed in assessment</div>
            <span class="exec-kpi-badge good">Active</span>
        </div>""", unsafe_allow_html=True)
    with k3:
        st.markdown(f"""
        <div class="exec-kpi-card amber">
            <span class="exec-kpi-icon">📋</span>
            <div class="exec-kpi-label">Number of DQ Rules</div>
            <div class="exec-kpi-value amber">{n_rules}</div>
            <div class="exec-kpi-delta">Validation rules executed</div>
            <span class="exec-kpi-badge warn">Rules</span>
        </div>""", unsafe_allow_html=True)
    with k4:
        st.markdown(f"""
        <div class="exec-kpi-card magenta">
            <span class="exec-kpi-icon">⭐</span>
            <div class="exec-kpi-label">Average DQ Score</div>
            <div class="exec-kpi-value magenta">{avg_score:.1f}%</div>
            <div class="exec-kpi-delta">Overall data quality score</div>
            <span class="exec-kpi-badge {'good' if avg_score>=80 else 'warn' if avg_score>=60 else 'danger'}">{badge_cls.title()}</span>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Dimension Gauges ──────────────────────────────────────────────────
    st.markdown("""
    <div class="dash-section-header">
        <div class="dash-section-dot magenta"></div>
        <h3>Dimension Scores</h3>
        <div class="dash-section-accent"></div>
    </div>
    """, unsafe_allow_html=True)

    dims = list(dim_scores.items()) if dim_scores else []
    # Pad to at least 4 gauges with 0
    gauge_labels = ["Completeness", "Standardization", "Uniqueness", "Validation"]
    gauge_vals   = {gl: dim_scores.get(gl, 0.0) for gl in gauge_labels}
    # Also include any extra dims
    for d, v in dim_scores.items():
        if d not in gauge_vals:
            gauge_vals[d] = v

    gauge_keys = list(gauge_vals.keys())[:4]
    gcols = st.columns(len(gauge_keys))
    for idx, gk in enumerate(gauge_keys):
        with gcols[idx]:
            st.markdown(
                f'<div class="dash-panel" style="text-align:center;padding:1rem;">'
                + _svg_gauge(gauge_vals[gk], gk)
                + f'<div class="dq-score-badge {("excellent" if gauge_vals[gk]>=80 else "good" if gauge_vals[gk]>=60 else "fair" if gauge_vals[gk]>=40 else "poor")}" style="margin:0.5rem auto 0;display:inline-flex;">'
                + f'{"✅" if gauge_vals[gk]>=80 else "⚠️" if gauge_vals[gk]>=60 else "❌"} {gauge_vals[gk]:.1f}%</div></div>',
                unsafe_allow_html=True,
            )

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Two-column: Stacked bar + Funnel ─────────────────────────────────
    st.markdown("""
    <div class="dash-section-header">
        <div class="dash-section-dot"></div>
        <h3>Distribution Analysis</h3>
        <div class="dash-section-accent"></div>
    </div>
    """, unsafe_allow_html=True)

    left, right = st.columns([3, 2])

    with left:
        # Overall DQ for individual master data → stacked horizontal bar
        st.markdown('<div class="dash-panel">', unsafe_allow_html=True)
        st.markdown('<div class="dash-panel-header"><span class="dash-panel-title">📊 Overall DQ by Dimension</span><span class="dash-panel-tag">Stacked</span></div>', unsafe_allow_html=True)

        all_dim_items = list((dim_scores or {}).items())
        if all_dim_items:
            total = sum(v for _, v in all_dim_items)
            bars_html = '<div class="stacked-bar-wrap">'
            for dname, dval in all_dim_items:
                pct_pass = min(100, dval)
                pct_fail = 100 - pct_pass
                bars_html += f"""
                <div style="margin-bottom:0.6rem;">
                    <div class="stacked-bar-label">
                        <span>{dname}</span><span>{dval:.1f}%</span>
                    </div>
                    <div class="stacked-bar-track">
                        <div class="stacked-bar-segment pass" style="width:{pct_pass:.1f}%">{pct_pass:.0f}%</div>
                        <div class="stacked-bar-segment fail" style="width:{pct_fail:.1f}%;background:#ede8f7;color:#9a85b8;">{pct_fail:.0f}%</div>
                    </div>
                </div>"""
            bars_html += "</div>"
            st.markdown(bars_html, unsafe_allow_html=True)
        else:
            st.info("No dimension scores available.")
        st.markdown('</div>', unsafe_allow_html=True)

    with right:
        # Funnel: Rules by dimension
        st.markdown('<div class="dash-panel">', unsafe_allow_html=True)
        st.markdown('<div class="dash-panel-header"><span class="dash-panel-title">🔻 DQ Rules by Dimension</span><span class="dash-panel-tag">Funnel</span></div>', unsafe_allow_html=True)

        if dim_scores and results_df is not None:
            dim_rule_counts = {}
            if "Dimension" in results_df.columns:
                dim_rule_counts = results_df["Dimension"].value_counts().to_dict()
            elif "dimension" in results_df.columns:
                dim_rule_counts = results_df["dimension"].value_counts().to_dict()
            else:
                # fallback: equal counts per dim
                dim_rule_counts = {d: max(1, int(len(results_df.columns) / max(len(dim_scores), 1))) for d in dim_scores}

            max_count = max(dim_rule_counts.values(), default=1)
            funnel_html = '<div class="funnel-wrap">'
            for i, (dname, cnt) in enumerate(sorted(dim_rule_counts.items(), key=lambda x: -x[1])):
                bar_pct = int(cnt / max_count * 100)
                cls = "magenta" if i % 2 == 1 else ""
                funnel_html += f"""
                <div class="funnel-bar">
                    <div class="funnel-bar-label">{dname}</div>
                    <div class="funnel-bar-track">
                        <div class="funnel-bar-fill {cls}" style="width:{bar_pct}%;">
                            <span class="funnel-bar-val">{cnt}</span>
                        </div>
                    </div>
                    <div class="funnel-bar-num">{cnt}</div>
                </div>"""
            funnel_html += "</div>"
            st.markdown(funnel_html, unsafe_allow_html=True)
        else:
            st.info("Run assessment to see rule distribution.")
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Column score detail table ─────────────────────────────────────────
    st.markdown("""
    <div class="dash-section-header">
        <div class="dash-section-dot magenta"></div>
        <h3>Column-level Score Detail</h3>
        <div class="dash-section-accent"></div>
    </div>
    """, unsafe_allow_html=True)

    if col_scores:
        sorted_cols = sorted(col_scores.items(), key=lambda x: x[1])
        rows_html = ""
        for cname, cscore in sorted_cols[:20]:
            cls  = "good" if cscore >= 80 else ("warn" if cscore >= 60 else "danger")
            icon = "✅" if cscore >= 80 else ("⚠️" if cscore >= 60 else "❌")
            rows_html += f"""
            <tr>
                <td>{cname}</td>
                <td><span class="score-pill {cls}">{icon} {cscore:.1f}%</span></td>
                <td>{"Passed" if cscore == 100 else "Failed"}</td>
            </tr>"""
        st.markdown(f"""
        <div class="dash-panel">
            <table class="score-table">
                <thead><tr><th>Column / CDE</th><th>DQ Score</th><th>Status</th></tr></thead>
                <tbody>{rows_html}</tbody>
            </table>
        </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
#  PAGE: HOME
# ══════════════════════════════════════════════════════════════════════════
def page_home():
    import time

    # ── Uniqus branded top bar ───────────────────────────────────────────
    render_uniqus_topbar("Data Quality Intelligence Studio")

    # ── Hero section ─────────────────────────────────────────────────────
    st.markdown("""
    <div class="uniqus-hero">
        <div class="uniqus-hero-badge">✦ Enterprise Data Governance Platform</div>
        <h1>Data Quality Intelligence Studio</h1>
        <p>Profile, validate, and monitor enterprise data using automated rules,
        dimension-based scoring, and executive-grade reporting — powered by Uniqus Consultech.</p>
    </div>
    """, unsafe_allow_html=True)

    # ── DQ Completion Banner ──────────────────────────────────────────────
    if st.session_state.dq_score is not None:
        sc  = st.session_state.dq_score
        lvl = dq_score_to_maturity_level(sc)
        badge_cls = "excellent" if sc >= 80 else ("good" if sc >= 60 else ("fair" if sc >= 40 else "poor"))

        st.markdown(f"""
        <div class="quick-stat-bar">
            <div class="quick-stat-item">
                <div class="quick-stat-val">{sc:.1f}%</div>
                <div class="quick-stat-lbl">DQ Score</div>
            </div>
            <div class="quick-stat-item">
                <div class="quick-stat-val magenta">{lvl}</div>
                <div class="quick-stat-lbl">Maturity Level</div>
            </div>
            <div class="quick-stat-item">
                <div class="quick-stat-val teal">{st.session_state.dq_object_name}</div>
                <div class="quick-stat-lbl">Master Object</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        col1, col2 = st.columns([4, 1])
        with col2:
            if st.button("View Results →", use_container_width=True):
                st.session_state["page"] = "dq"
                st.rerun()
        st.markdown("<br>", unsafe_allow_html=True)

    # ── Solutions Workspace cards ─────────────────────────────────────────
    st.markdown("""
    <div class="dash-section-header">
        <div class="dash-section-dot"></div>
        <h3>Solutions Workspace</h3>
        <div class="dash-section-accent"></div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns(2, gap="large")

    with col1:
        st.markdown("""
        <div class="nav-card purple">
            <div class="nav-card-icon-wrap">🔍</div>
            <div class="nav-card-title">Data Quality Assessment</div>
            <div class="nav-card-desc">Upload dataset and rules to generate automated DQ scores,
            column analysis, dimension scoring and enterprise reports.</div>
            <span class="nav-card-arrow">→</span>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Start DQ Assessment →", use_container_width=True, key="home_dq"):
            st.session_state["page"] = "dq"
            st.rerun()

    with col2:
        st.markdown("""
        <div class="nav-card magenta">
            <div class="nav-card-icon-wrap">📈</div>
            <div class="nav-card-title">Data Maturity Assessment</div>
            <div class="nav-card-desc">Evaluate DAMA maturity dimensions, generate executive visuals,
            PDF reports and Excel outputs.</div>
            <span class="nav-card-arrow">→</span>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Start Maturity Assessment →", use_container_width=True, key="home_mat"):
            st.session_state["page"] = "maturity"
            st.rerun()

    st.markdown("<br>", unsafe_allow_html=True)
    col3, col4 = st.columns(2, gap="large")

    with col3:
        st.markdown("""
        <div class="nav-card teal">
            <div class="nav-card-icon-wrap">📋</div>
            <div class="nav-card-title">Policy Hub</div>
            <div class="nav-card-desc">Central governance repository for policy workflows,
            approval tracking and compliance monitoring.</div>
            <span class="nav-card-arrow">→</span>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Open Policy Hub →", use_container_width=True, key="home_policy"):
            st.session_state["page"] = "policy"
            st.rerun()

    with col4:
        st.markdown("""
        <div class="nav-card amber">
            <div class="nav-card-icon-wrap">🎯</div>
            <div class="nav-card-title">Case Management</div>
            <div class="nav-card-desc">Track and resolve data quality issues with ownership,
            SLA tracking and audit trails.</div>
            <span class="nav-card-arrow">→</span>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Open Case Management →", use_container_width=True, key="home_cases"):
            st.session_state["page"] = "cases"
            st.rerun()

    st.divider()
# ══════════════════════════════════════════════════════════════════════════
#  PDF REPORT GENERATOR  — Power BI style static dashboard PDF
# ══════════════════════════════════════════════════════════════════════════
def _build_dq_pdf_report(
    overall: float,
    dim_scores: dict,
    results_df,
    col_scores: dict,
    obj_name: str = "Dataset",
) -> bytes:
    """
    Generate a multi-page PDF report that mirrors the PBIX layout:
    Page 1 — Executive Outlook
    Page N — One page per dimension (Completeness, Standardization, etc.)
    Returns raw PDF bytes.
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    from matplotlib.patches import FancyBboxPatch
    from io import BytesIO

    PURPLE  = "#5b2d90"
    MAGENTA = "#b10f74"
    AMBER   = "#d97706"
    TEAL    = "#0d9488"
    GREEN   = "#10b981"
    GREY_BG = "#f0edf8"
    CARD_BG = "#ffffff"
    TEXT1   = "#1a1a2e"
    TEXT2   = "#4a4a6a"
    TEXT3   = "#7a7a9a"

    DIM_COLORS = {
        "Completeness":    AMBER,
        "Standardization": AMBER,
        "Uniqueness":      TEAL,
        "Validation":      AMBER,
    }

    def _score_color(s):
        if s >= 90: return TEAL
        if s >= 75: return AMBER
        return MAGENTA

    def _draw_gauge(ax, score, color, label, fontsize=18):
        """Draw a half-donut gauge on an axes."""
        ax.set_xlim(-1.1, 1.1); ax.set_ylim(-0.15, 1.1); ax.axis("off")
        # background arc
        theta = np.linspace(np.pi, 0, 200)
        r_outer, r_inner = 1.0, 0.62
        x_bg = np.concatenate([r_outer*np.cos(theta), r_inner*np.cos(theta[::-1])])
        y_bg = np.concatenate([r_outer*np.sin(theta), r_inner*np.sin(theta[::-1])])
        ax.fill(x_bg, y_bg, color="#e5e7eb", zorder=1)
        # filled arc
        end_ang = np.pi - (score / 100) * np.pi
        theta2  = np.linspace(np.pi, end_ang, 200)
        x_fg    = np.concatenate([r_outer*np.cos(theta2), r_inner*np.cos(theta2[::-1])])
        y_fg    = np.concatenate([r_outer*np.sin(theta2), r_inner*np.sin(theta2[::-1])])
        ax.fill(x_fg, y_fg, color=color, zorder=2)
        # centre text
        ax.text(0, 0.22, f"{score:.2f}%", ha="center", va="center",
                fontsize=fontsize, fontweight="bold", color=TEXT2, zorder=3)
        # limits
        ax.text(-0.95, -0.12, "0.00%", ha="left", va="top",
                fontsize=6.5, color=TEXT3)
        ax.text(0.95, -0.12, "100.00%", ha="right", va="top",
                fontsize=6.5, color=TEXT3)
        # label above
        ax.text(0, 1.05, label, ha="center", va="bottom",
                fontsize=8, fontweight="bold", color=TEXT2)

    def _kpi_card(fig, rect, title, value, sub=None):
        """Draw a KPI card (rounded rect) at given [left,bottom,width,height]."""
        ax = fig.add_axes(rect)
        ax.set_xlim(0, 1); ax.set_ylim(0, 1); ax.axis("off")
        fp = FancyBboxPatch((0.02, 0.05), 0.96, 0.90,
                            boxstyle="round,pad=0.03",
                            linewidth=1.5, edgecolor="#d9cef0",
                            facecolor=CARD_BG, zorder=0)
        ax.add_patch(fp)
        ax.text(0.5, 0.82, title, ha="center", va="top",
                fontsize=7.5, fontweight="bold", color=TEXT2)
        ax.text(0.5, 0.45, str(value), ha="center", va="center",
                fontsize=16, fontweight="black", color=PURPLE)
        if sub:
            for i, (k, v) in enumerate(sub.items()):
                ax.text(0.15 + i * 0.35, 0.2, str(v), ha="center",
                        fontsize=9, fontweight="bold", color=PURPLE)
                ax.text(0.15 + i * 0.35, 0.08, k, ha="center",
                        fontsize=6.5, color=TEXT3)

    def _page_header(fig, title):
        ax = fig.add_axes([0, 0.93, 1.0, 0.07])
        ax.set_xlim(0, 1); ax.set_ylim(0, 1); ax.axis("off")
        ax.add_patch(plt.Rectangle((0, 0), 1, 1, color=PURPLE))
        ax.text(0.5, 0.5, title, ha="center", va="center",
                fontsize=16, fontweight="bold", color="white")

    def _stacked_bar_chart(ax, data, title):
        """data = {master: {High:%, Low:%, Medium:%}}"""
        ax.set_title(title, fontsize=8, fontweight="bold", color=TEXT2, pad=4)
        masters = list(data.keys())
        highs   = [data[m].get("High", 0)   for m in masters]
        mediums = [data[m].get("Medium", 0) for m in masters]
        lows    = [data[m].get("Low", 0)    for m in masters]
        x = range(len(masters))
        ax.bar(x, highs,   color=TEAL,    label="High",   width=0.5)
        ax.bar(x, mediums, bottom=highs,  color=AMBER,    label="Medium", width=0.5)
        bot2 = [h+m for h, m in zip(highs, mediums)]
        ax.bar(x, lows, bottom=bot2,      color=MAGENTA,  label="Low",    width=0.5)
        ax.set_xticks(list(x)); ax.set_xticklabels(masters, fontsize=7.5)
        ax.yaxis.set_visible(False); ax.spines[:].set_visible(False)
        ax.legend(fontsize=6, loc="upper right", frameon=False)
        for xi, (h, m, l) in enumerate(zip(highs, mediums, lows)):
            if h > 5: ax.text(xi, h/2,        f"{h:.1f}%", ha="center", va="center", fontsize=6.5, color="white", fontweight="bold")
            if m > 5: ax.text(xi, h + m/2,     f"{m:.1f}%", ha="center", va="center", fontsize=6.5, color="white", fontweight="bold")
            if l > 5: ax.text(xi, h + m + l/2, f"{l:.1f}%", ha="center", va="center", fontsize=6.5, color="white", fontweight="bold")

    def _horiz_bar(ax, items, title, color=AMBER):
        ax.set_title(title, fontsize=8, fontweight="bold", color=TEXT2, pad=4)
        names  = [i[0] for i in items]
        values = [i[1] for i in items]
        ypos   = range(len(names))
        ax.barh(list(ypos), values, color=color, height=0.45)
        ax.set_yticks(list(ypos)); ax.set_yticklabels(names, fontsize=7.5)
        ax.set_xlim(0, max(values)*1.18 if values else 10)
        ax.spines[:].set_visible(False); ax.xaxis.set_visible(False)
        for y, v in zip(ypos, values):
            ax.text(v + max(values)*0.02 if values else 0.5, y,
                    f"{v:.2f}%" if isinstance(v, float) else str(v),
                    va="center", fontsize=7, fontweight="bold", color=TEXT2)

    def _detail_table(ax, rows, cols_header):
        ax.axis("off")
        if not rows:
            ax.text(0.5, 0.5, "No data", ha="center", fontsize=8, color=TEXT3)
            return
        table_data = [cols_header] + rows[:12]
        t = ax.table(cellText=table_data[1:], colLabels=table_data[0],
                     cellLoc="left", loc="upper left", bbox=[0, 0, 1, 1])
        t.auto_set_font_size(False)
        t.set_fontsize(6.5)
        for (r, c), cell in t.get_celld().items():
            cell.set_edgecolor("#d9cef0")
            if r == 0:
                cell.set_facecolor(PURPLE)
                cell.set_text_props(color="white", fontweight="bold")
            elif r % 2 == 0:
                cell.set_facecolor("#f5f0fc")
            else:
                cell.set_facecolor(CARD_BG)

    pdf_pages = BytesIO()

    # ── derive stats ──────────────────────────────────────────────────────
    n_cols    = len(col_scores) if col_scores else 0
    n_records = len(results_df) if results_df is not None else 0
    n_rules   = 0
    if results_df is not None and "rule" in [c.lower() for c in results_df.columns]:
        n_rules = results_df.shape[0]
    else:
        n_rules = max(len(col_scores or {}), 1)

    dims_available = list((dim_scores or {}).keys())

    # ── helper: stacked chart data from results ───────────────────────────
    def _stacked_data_for_dim(dim_name):
        """Build {master_or_overall: {High/Medium/Low: pct}} from col_scores."""
        if col_scores:
            high_pct   = sum(1 for s in col_scores.values() if s >= 90) / max(len(col_scores), 1) * 100
            medium_pct = sum(1 for s in col_scores.values() if 60 <= s < 90) / max(len(col_scores), 1) * 100
            low_pct    = sum(1 for s in col_scores.values() if s < 60) / max(len(col_scores), 1) * 100
            return {obj_name: {"High": round(high_pct,1), "Medium": round(medium_pct,1), "Low": round(low_pct,1)}}
        return {}

    def _avg_score_data():
        """Build [(name, score)] for avg score chart."""
        return [(k, round(v, 2)) for k, v in sorted((col_scores or {}).items(), key=lambda x: -x[1])[:8]]

    def _rules_by_master():
        """Build [(name, count)] funnel data."""
        if results_df is not None and "dimension" in [c.lower() for c in results_df.columns]:
            col = [c for c in results_df.columns if c.lower() == "dimension"][0]
            counts = results_df[col].value_counts().to_dict()
            return sorted(counts.items(), key=lambda x: -x[1])[:6]
        return [(d, max(1, int(n_rules/max(len(dims_available),1)))) for d in dims_available]

    def _detail_rows_for_dim(dim_name):
        """Return top rows for detail table for a given dimension."""
        rows = []
        for cname, cscore in sorted((col_scores or {}).items(), key=lambda x: -x[1])[:12]:
            thresh = "High" if cscore >= 80 else ("Medium" if cscore >= 60 else "Low")
            non_c  = 0
            if results_df is not None and "Count of issues" in results_df.columns:
                mask = results_df["Count of issues"].fillna(0)
                non_c = int(mask.mean()) if len(mask) else 0
            rows.append([obj_name, dim_name, cname, f"{cscore:.2f}%", str(non_c), thresh])
        return rows

    from matplotlib.backends.backend_pdf import PdfPages

    with PdfPages(pdf_pages) as pdf:

        # ══════════════════════════════════════════════════════════════════
        # PAGE 1 — EXECUTIVE OUTLOOK
        # ══════════════════════════════════════════════════════════════════
        fig = plt.figure(figsize=(16, 9), facecolor=GREY_BG)

        # header bar
        _page_header(fig, f"Data Quality Assessment - Executive Outlook  |  {obj_name}")

        # ── 4 KPI cards ──────────────────────────────────────────────────
        kpi_y    = 0.78
        kpi_h    = 0.13
        kpi_w    = 0.22
        kpi_gap  = 0.025
        kpi_start = 0.02
        _kpi_card(fig, [kpi_start,              kpi_y, kpi_w, kpi_h], "Number of CDEs",              n_cols)
        _kpi_card(fig, [kpi_start+kpi_w+kpi_gap,kpi_y, kpi_w, kpi_h], "Number of Active Records",   n_records,
                  sub={obj_name: n_records})
        _kpi_card(fig, [kpi_start+2*(kpi_w+kpi_gap),kpi_y, kpi_w, kpi_h], "Number of Data Quality Rules", n_rules)
        _kpi_card(fig, [kpi_start+3*(kpi_w+kpi_gap),kpi_y, kpi_w, kpi_h], "Average Data Quality Score",
                  f"{overall:.2f}%")

        # ── 4 Gauge charts ──────────────────────────────────────────────
        gauge_dims = ["Completeness", "Standardization", "Uniqueness", "Validation"]
        for gi, gd in enumerate(gauge_dims):
            gscore = dim_scores.get(gd, overall) if dim_scores else overall
            gcol   = DIM_COLORS.get(gd, PURPLE)
            if gd == "Uniqueness": gcol = TEAL
            gax = fig.add_axes([0.02 + gi*0.245, 0.48, 0.22, 0.28])
            gax.set_facecolor(CARD_BG)
            fp  = FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.01",
                                 linewidth=1, edgecolor="#d9cef0", facecolor=CARD_BG,
                                 transform=gax.transAxes, zorder=-1)
            gax.add_patch(fp)
            _draw_gauge(gax, gscore, gcol, gd, fontsize=16)

        # ── Bottom row: stacked bar + bar chart + donut ──────────────────
        # Stacked bar: Overall DQ per master data
        ax_stacked = fig.add_axes([0.02, 0.05, 0.30, 0.40])
        ax_stacked.set_facecolor(CARD_BG)
        stacked_data = _stacked_data_for_dim("overall")
        if stacked_data:
            _stacked_bar_chart(ax_stacked, stacked_data, "Overall DQ for Individual Master Data")
        else:
            ax_stacked.text(0.5, 0.5, "No master data split available",
                            ha="center", fontsize=8, color=TEXT3, transform=ax_stacked.transAxes)
            ax_stacked.set_title("Overall DQ for Individual Master Data",
                                  fontsize=8, fontweight="bold", color=TEXT2, pad=4)

        # Avg DQ score bar chart
        ax_bar = fig.add_axes([0.34, 0.05, 0.32, 0.40])
        ax_bar.set_facecolor(CARD_BG)
        score_data = _avg_score_data()[:5]
        if score_data:
            _horiz_bar(ax_bar, [(n, s) for n, s in score_data],
                       "Distribution of DQ Score by Column")

        # Donut — dimension distribution
        ax_donut = fig.add_axes([0.67, 0.05, 0.30, 0.40])
        ax_donut.set_facecolor(CARD_BG)
        if dim_scores:
            wedge_sizes  = list(dim_scores.values())
            wedge_labels = list(dim_scores.keys())
            wedge_colors = [PURPLE, MAGENTA, TEAL, AMBER, "#a78bfa", "#34d399"][:len(wedge_sizes)]
            wedges, texts = ax_donut.pie(
                wedge_sizes, labels=None, colors=wedge_colors,
                startangle=90, wedgeprops=dict(width=0.55, edgecolor="white", linewidth=2),
            )
            ax_donut.legend(wedge_labels, loc="lower center", fontsize=6.5,
                            frameon=False, ncol=2, bbox_to_anchor=(0.5, -0.12))
            ax_donut.set_title("Distribution of CDEs by Dimension", fontsize=8,
                               fontweight="bold", color=TEXT2, pad=4)
        else:
            ax_donut.text(0.5, 0.5, "—", ha="center", fontsize=12, color=TEXT3)

        # page nav tabs strip
        ax_tabs = fig.add_axes([0, 0, 1, 0.04])
        ax_tabs.set_xlim(0, 1); ax_tabs.set_ylim(0, 1); ax_tabs.axis("off")
        ax_tabs.add_patch(plt.Rectangle((0, 0), 1, 1, color="#2d1b50"))
        tab_labels = ["DQA - Executive Outlook"] + [f"DQA - {d}" for d in gauge_dims]
        for ti, tl in enumerate(tab_labels):
            tx   = 0.02 + ti * 0.19
            bg   = PURPLE if ti == 0 else "#2d1b50"
            fc   = "white"
            ax_tabs.add_patch(plt.Rectangle((tx, 0.05), 0.18, 0.9,
                                            color=bg, zorder=1))
            ax_tabs.text(tx + 0.09, 0.5, tl, ha="center", va="center",
                         fontsize=7.5, color=fc, fontweight="bold" if ti == 0 else "normal",
                         zorder=2)

        plt.tight_layout(rect=[0, 0.04, 1, 0.93], h_pad=0.3, w_pad=0.3)
        pdf.savefig(fig, bbox_inches="tight", facecolor=GREY_BG)
        plt.close(fig)

        # ══════════════════════════════════════════════════════════════════
        # PAGES 2–5 — ONE PER DIMENSION
        # ══════════════════════════════════════════════════════════════════
        all_dims = gauge_dims if dim_scores else dims_available
        for page_idx, dim_name in enumerate(all_dims):
            dim_score = (dim_scores or {}).get(dim_name, overall)
            dim_color = DIM_COLORS.get(dim_name, PURPLE)
            if dim_name == "Uniqueness": dim_color = TEAL

            fig2 = plt.figure(figsize=(16, 9), facecolor=GREY_BG)
            _page_header(fig2, f"Data Quality Assessment - {dim_name}  |  {obj_name}")

            # 4 KPI cards (same layout each page)
            _kpi_card(fig2, [kpi_start,               kpi_y, kpi_w, kpi_h], "Number of CDEs",           n_cols)
            _kpi_card(fig2, [kpi_start+kpi_w+kpi_gap, kpi_y, kpi_w, kpi_h], "Number of Datasets",      max(1, n_cols//8))
            _kpi_card(fig2, [kpi_start+2*(kpi_w+kpi_gap), kpi_y, kpi_w, kpi_h], "Number of DQ Rules",  n_rules)
            _kpi_card(fig2, [kpi_start+3*(kpi_w+kpi_gap), kpi_y, kpi_w, kpi_h], "Average DQ Score",    f"{dim_score:.2f}%")

            # Gauge (left)
            gax2 = fig2.add_axes([0.02, 0.36, 0.22, 0.36])
            gax2.set_facecolor(CARD_BG)
            fp2  = FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.01",
                                  linewidth=1, edgecolor="#d9cef0", facecolor=CARD_BG,
                                  transform=gax2.transAxes, zorder=-1)
            gax2.add_patch(fp2)
            _draw_gauge(gax2, dim_score, dim_color, f"{dim_name} Score", fontsize=18)

            # Stacked bar (middle)
            ax_stk2 = fig2.add_axes([0.26, 0.36, 0.40, 0.36])
            ax_stk2.set_facecolor(CARD_BG)
            stacked2 = _stacked_data_for_dim(dim_name)
            if stacked2:
                _stacked_bar_chart(ax_stk2, stacked2,
                                   "Number of DQ Rules Across Threshold Categories")
            else:
                ax_stk2.set_title("Number of DQ Rules Across Threshold Categories",
                                   fontsize=8, fontweight="bold", color=TEXT2, pad=4)
                ax_stk2.axis("off")

            # Rules by master funnel (right)
            ax_funnel = fig2.add_axes([0.69, 0.36, 0.29, 0.36])
            ax_funnel.set_facecolor(CARD_BG)
            rules_data = _rules_by_master()[:5]
            if rules_data:
                _horiz_bar(ax_funnel, [(str(n), v) for n, v in rules_data],
                           "Number of DQ Rules by Dimension", color=dim_color)

            # Detail table (bottom left / centre)
            ax_tbl = fig2.add_axes([0.02, 0.04, 0.64, 0.30])
            ax_tbl.set_facecolor(CARD_BG)
            fp3  = FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.01",
                                  linewidth=1, edgecolor="#d9cef0", facecolor=CARD_BG,
                                  transform=ax_tbl.transAxes, zorder=-1)
            ax_tbl.add_patch(fp3)
            detail_rows = _detail_rows_for_dim(dim_name)
            _detail_table(ax_tbl, detail_rows,
                          ["System", "Domain", "Field Name", "DQ Score", "Non-Compliant", "Threshold"])

            # Avg DQ score by column (bottom right)
            ax_avg = fig2.add_axes([0.69, 0.04, 0.29, 0.30])
            ax_avg.set_facecolor(CARD_BG)
            avg_items = [(k, round(v, 2)) for k, v in sorted(
                (col_scores or {}).items(), key=lambda x: -x[1])[:6]]
            if avg_items:
                _horiz_bar(ax_avg, avg_items,
                           "Average DQ Score by Column", color=dim_color)

            # Tab nav
            ax_tabs2 = fig2.add_axes([0, 0, 1, 0.04])
            ax_tabs2.set_xlim(0, 1); ax_tabs2.set_ylim(0, 1); ax_tabs2.axis("off")
            ax_tabs2.add_patch(plt.Rectangle((0, 0), 1, 1, color="#2d1b50"))
            for ti, tl in enumerate(tab_labels):
                tx   = 0.02 + ti * 0.19
                is_active = (tl.replace("DQA - ", "") == dim_name)
                bg   = PURPLE if is_active else "#2d1b50"
                ax_tabs2.add_patch(plt.Rectangle((tx, 0.05), 0.18, 0.9, color=bg, zorder=1))
                ax_tabs2.text(tx + 0.09, 0.5, tl, ha="center", va="center",
                              fontsize=7.5, color="white",
                              fontweight="bold" if is_active else "normal", zorder=2)

            plt.tight_layout(rect=[0, 0.04, 1, 0.93], h_pad=0.3, w_pad=0.3)
            pdf.savefig(fig2, bbox_inches="tight", facecolor=GREY_BG)
            plt.close(fig2)

    pdf_pages.seek(0)
    return pdf_pages.read()


# ══════════════════════════════════════════════════════════════════════════
#  PAGE: DQ ASSESSMENT
# ══════════════════════════════════════════════════════════════════════════
def _render_dq_results(overall, dim_scores, results, col_scores, obj_name, pdf_bytes, xl_path, excel_filename, rb_path_str):
    """
    Render the full DQ results dashboard.
    Called both immediately after a run AND when returning to the DQ page from session state.
    """
    lvl = dq_score_to_maturity_level(overall)

    # ── Score summary bar ────────────────────────────────────────────────
    st.markdown(f"""
    <div class="quick-stat-bar" style="margin:1.5rem 0;">
        <div class="quick-stat-item">
            <div class="quick-stat-val">{overall:.1f}%</div>
            <div class="quick-stat-lbl">Overall DQ Score</div>
        </div>
        <div class="quick-stat-item">
            <div class="quick-stat-val magenta">{lvl}</div>
            <div class="quick-stat-lbl">Maturity Level</div>
        </div>
        <div class="quick-stat-item">
            <div class="quick-stat-val teal">{len(col_scores) if col_scores else 0}</div>
            <div class="quick-stat-lbl">Columns Analysed</div>
        </div>
        <div class="quick-stat-item">
            <div class="quick-stat-val" style="color:#d97706;">{len(results):,}</div>
            <div class="quick-stat-lbl">Records Processed</div>
        </div>
    </div>
    <div class="dq-result-session-note">
        <span>📌</span>
        <span>Results are saved for this session. Upload new files above to run a fresh assessment.</span>
    </div>
    """, unsafe_allow_html=True)

    # ── Executive Dashboard ──────────────────────────────────────────────
    render_static_dq_dashboard(overall, dim_scores, results, col_scores)

    # ── Dimension detail tabs ────────────────────────────────────────────
    st.markdown("""
    <div class="dash-section-header" style="margin-top:1.5rem;">
        <div class="dash-section-dot magenta"></div>
        <h3>Dimension Detail Analysis</h3>
        <div class="dash-section-accent"></div>
    </div>
    """, unsafe_allow_html=True)

    dim_tab_names = list(dim_scores.keys()) if dim_scores else ["Results"]
    dim_tabs = st.tabs(dim_tab_names)
    for dti, dtname in enumerate(dim_tab_names):
        with dim_tabs[dti]:
            dscore  = dim_scores.get(dtname, overall)
            d_clean = len([c for c, s in col_scores.items() if s >= 80]) if col_scores else 0
            d_fail  = len(col_scores) - d_clean if col_scores else 0
            st.markdown(f"""
            <div class="quick-stat-bar">
                <div class="quick-stat-item">
                    <div class="quick-stat-val">{dscore:.1f}%</div>
                    <div class="quick-stat-lbl">{dtname} Score</div>
                </div>
                <div class="quick-stat-item">
                    <div class="quick-stat-val teal">{d_clean}</div>
                    <div class="quick-stat-lbl">Columns ≥80%</div>
                </div>
                <div class="quick-stat-item">
                    <div class="quick-stat-val magenta">{d_fail}</div>
                    <div class="quick-stat-lbl">Needs Attention</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            if col_scores:
                sorted_scores = sorted(col_scores.items(), key=lambda x: x[1])[:20]
                rows_html = ""
                for cname, cscore in sorted_scores:
                    cls    = "good"   if cscore >= 80 else ("warn" if cscore >= 60 else "danger")
                    icon   = "✅"     if cscore >= 80 else ("⚠️"    if cscore >= 60 else "❌")
                    thresh = "High"   if cscore >= 80 else ("Medium" if cscore >= 60 else "Low")
                    non_c  = 0
                    if results is not None and "Count of issues" in results.columns:
                        non_c = int(results["Count of issues"].fillna(0).mean())
                    rows_html += f"""
                    <tr>
                        <td>{obj_name}</td><td>{dtname}</td><td>{cname}</td>
                        <td><span class="score-pill {cls}">{icon} {cscore:.2f}%</span></td>
                        <td>{non_c}</td>
                        <td><span class="score-pill {cls}">{thresh}</span></td>
                    </tr>"""
                st.markdown(f"""
                <div class="dash-panel">
                    <table class="score-table">
                        <thead><tr>
                            <th>Domain</th><th>Dimension</th><th>Field Name</th>
                            <th>DQ Score</th><th>Non-Compliant</th><th>Threshold</th>
                        </tr></thead>
                        <tbody>{rows_html}</tbody>
                    </table>
                </div>
                """, unsafe_allow_html=True)

    # ── Download section ─────────────────────────────────────────────────
    st.markdown("""
    <div class="dash-section-header" style="margin-top:1.5rem;">
        <div class="dash-section-dot"></div>
        <h3>Download Reports</h3>
        <div class="dash-section-accent"></div>
    </div>
    """, unsafe_allow_html=True)

    d1, d2, d3 = st.columns(3)
    with d1:
        st.markdown("""
        <div class="dq-dl-card">
            <div class="dq-dl-icon">📄</div>
            <div class="dq-dl-title">PDF Report</div>
            <div class="dq-dl-desc">Power BI–style executive dashboard with all 5 pages</div>
        </div>
        """, unsafe_allow_html=True)
        if pdf_bytes:
            pdf_fname = get_timestamp_filename(f"DQ_Assessment_{obj_name or 'Report'}", "pdf")
            st.download_button(
                "⬇ Download PDF Report", data=pdf_bytes,
                file_name=pdf_fname, mime="application/pdf",
                use_container_width=True, key="dq_pdf_dl",
            )
    with d2:
        st.markdown("""
        <div class="dq-dl-card">
            <div class="dq-dl-icon">📊</div>
            <div class="dq-dl-title">Excel Report</div>
            <div class="dq-dl-desc">Full results, scores, dimension analysis &amp; annexures</div>
        </div>
        """, unsafe_allow_html=True)
        if xl_path and Path(xl_path).exists():
            with open(xl_path, "rb") as f:
                st.download_button(
                    "⬇ Download Excel Report", data=f.read(),
                    file_name=excel_filename or "DQ_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="dq_xl_dl",
                )
    with d3:
        st.markdown("""
        <div class="dq-dl-card">
            <div class="dq-dl-icon">🔗</div>
            <div class="dq-dl-title">Rulebook JSON</div>
            <div class="dq-dl-desc">Generated rule configuration for reuse and audit</div>
        </div>
        """, unsafe_allow_html=True)
        if rb_path_str and Path(rb_path_str).exists():
            with open(rb_path_str, "rb") as f:
                rb_fn = get_timestamp_filename("Rulebook", "json")
                st.download_button(
                    "⬇ Download Rulebook", data=f.read(),
                    file_name=rb_fn, mime="application/json",
                    use_container_width=True, key="dq_rb_dl",
                )

    # ── Next step ────────────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    _, nc, _ = st.columns([1, 1.2, 1])
    with nc:
        if st.button("📈 Continue to Maturity Assessment →",
                     type="primary", use_container_width=True, key="dq_to_mat"):
            st.session_state["page"] = "maturity"; st.rerun()


def page_dq():
    with st.sidebar:
        st.markdown("### 🧭 Navigation")
        if st.button("🏠 Home",      use_container_width=True, key="dq_home"):
            st.session_state["page"] = "home"; st.rerun()
        if st.button("📈 Maturity",  use_container_width=True, key="dq_maturity"):
            st.session_state["page"] = "maturity"; st.rerun()
        if st.button("📋 Policies",  use_container_width=True, key="dq_policy"):
            st.session_state["page"] = "policy"; st.rerun()
        if st.button("🎯 Cases",     use_container_width=True, key="dq_cases"):
            st.session_state["page"] = "cases"; st.rerun()
        st.divider()

    # ── Top bar + Hero ────────────────────────────────────────────────────
    render_uniqus_topbar("Data Quality Assessment")

    st.markdown("""
    <div class="uniqus-hero" style="padding:2rem 2.5rem 1.8rem;">
        <div class="uniqus-hero-badge">🔍 Enterprise DQ Engine</div>
        <h1 style="font-size:2rem!important;">Data Quality Assessment</h1>
        <p style="margin-bottom:0;">Upload your master dataset and rules configuration to generate
        automated dimension scores, column analysis, and Power BI–style executive reports.</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Upload zone (always shown) ────────────────────────────────────────
    st.markdown("""
    <div class="dash-section-header">
        <div class="dash-section-dot"></div>
        <h3>Input Files</h3>
        <div class="dash-section-accent"></div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns(2, gap="large")
    with col1:
        st.markdown("""
        <div class="dq-upload-zone">
            <div class="dq-upload-icon">📂</div>
            <div class="dq-upload-title">Master Dataset</div>
            <div class="dq-upload-hint">CSV · Excel (.xlsx/.xls/.xlsm) · JSON · Parquet · ODS · XML</div>
        </div>
        """, unsafe_allow_html=True)
        data_file = st.file_uploader(
            "Upload Master Dataset",
            type=AppConfig.SUPPORTED_DATA_FORMATS,
            label_visibility="collapsed",
            key="dq_data_uploader",
        )

    with col2:
        st.markdown("""
        <div class="dq-upload-zone">
            <div class="dq-upload-icon">📋</div>
            <div class="dq-upload-title">Rules Configuration</div>
            <div class="dq-upload-hint">CSV/Excel: column_name · rule · dimension · message &nbsp;|&nbsp; JSON rulebook</div>
        </div>
        """, unsafe_allow_html=True)
        rules_file = st.file_uploader(
            "Upload Rules File",
            type=AppConfig.SUPPORTED_RULES_FORMATS + ["json"],
            label_visibility="collapsed",
            key="dq_rules_uploader",
        )

    # ══════════════════════════════════════════════════════════════════════
    #  IF RESULTS ALREADY EXIST IN SESSION → show them immediately
    #  (persists across page navigation until a new file is uploaded)
    # ══════════════════════════════════════════════════════════════════════
    has_results = st.session_state.get("dq_score") is not None

    if has_results and not data_file:
        # Show persisted results — no new file uploaded yet
        _render_dq_results(
            overall       = st.session_state["dq_score"],
            dim_scores    = st.session_state.get("dq_dim_scores", {}),
            results       = st.session_state["dq_results_df"],
            col_scores    = st.session_state.get("dq_col_scores", {}),
            obj_name      = st.session_state.get("dq_object_name", "Dataset"),
            pdf_bytes     = st.session_state.get("dq_pdf_bytes"),
            xl_path       = st.session_state.get("dq_excel_path"),
            excel_filename= st.session_state.get("dq_excel_filename", "DQ_Report.xlsx"),
            rb_path_str   = st.session_state.get("dq_rb_path"),
        )
        return

    # ── Empty state (no results, no files) ───────────────────────────────
    if not data_file or not rules_file:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("""
        <div class="dq-steps-row">
            <div class="dq-step active">
                <div class="dq-step-num">01</div>
                <div class="dq-step-icon">📤</div>
                <div class="dq-step-title">Upload Files</div>
                <div class="dq-step-desc">Master dataset + rules configuration CSV or JSON</div>
            </div>
            <div class="dq-step-line"></div>
            <div class="dq-step">
                <div class="dq-step-num">02</div>
                <div class="dq-step-icon">⚙️</div>
                <div class="dq-step-title">Run Assessment</div>
                <div class="dq-step-desc">Automated rule execution across all dimensions</div>
            </div>
            <div class="dq-step-line"></div>
            <div class="dq-step">
                <div class="dq-step-num">03</div>
                <div class="dq-step-icon">📊</div>
                <div class="dq-step-title">Executive Dashboard</div>
                <div class="dq-step-desc">Power BI–style report + PDF download</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        f1, f2 = st.columns(2, gap="large")
        with f1:
            st.markdown("""
            <div class="dash-panel">
                <div class="dash-panel-header">
                    <span class="dash-panel-title">📂 Accepted Data Formats</span>
                </div>
                <div class="dq-format-grid">
                    <span class="dq-fmt-chip">CSV</span><span class="dq-fmt-chip">XLSX</span>
                    <span class="dq-fmt-chip">XLS</span><span class="dq-fmt-chip">XLSM</span>
                    <span class="dq-fmt-chip">JSON</span><span class="dq-fmt-chip">Parquet</span>
                    <span class="dq-fmt-chip">ODS</span><span class="dq-fmt-chip">XML</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
        with f2:
            st.markdown("""
            <div class="dash-panel">
                <div class="dash-panel-header">
                    <span class="dash-panel-title">📋 Rules File Required Columns</span>
                </div>
                <table class="score-table" style="margin-top:0.25rem;">
                    <thead><tr><th>Column</th><th>Description</th></tr></thead>
                    <tbody>
                        <tr><td><code>column_name</code></td><td>Target data column</td></tr>
                        <tr><td><code>rule</code></td><td>Validation rule type</td></tr>
                        <tr><td><code>dimension</code></td><td>DQ dimension (Completeness etc.)</td></tr>
                        <tr><td><code>message</code></td><td>Failure message</td></tr>
                    </tbody>
                </table>
            </div>
            """, unsafe_allow_html=True)
        return

    # ── Configuration ─────────────────────────────────────────────────────
    st.markdown("""
    <div class="dash-section-header" style="margin-top:1.5rem;">
        <div class="dash-section-dot magenta"></div>
        <h3>Assessment Configuration</h3>
        <div class="dash-section-accent"></div>
    </div>
    """, unsafe_allow_html=True)

    cfg1, cfg2 = st.columns(2)
    with cfg1:
        obj_name = st.text_input(
            "Master Object / Dataset Name",
            value=st.session_state.get("dq_object_name", "Customer"),
            placeholder="e.g. Customer, Vendor, Material…",
            help="Labels reports and links results to the Maturity Assessment.",
            key="dq_obj_name_input",
        )
    with cfg2:
        sheet_name = None
        if data_file.name.lower().endswith((".xlsx", ".xls", ".xlsm")):
            loader = FileLoaderService()
            tmp    = AppConfig.TEMP_DIR / data_file.name
            tmp.write_bytes(data_file.getbuffer())
            sheets = loader.get_sheet_names(tmp)
            if len(sheets) > 1:
                sheet_name = st.selectbox("Select Sheet", sheets, key="dq_sheet")

    # ── Run Button ────────────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    _, rc, _ = st.columns([1, 1.2, 1])
    with rc:
        run_button = st.button(
            "🚀 Run DQ Assessment",
            type="primary",
            use_container_width=True,
            key="dq_run",
        )

    if not run_button:
        return

    try:
        clean_temp_directory()

        # Progress bar with dark text so it's visible
        st.markdown("""
        <style>
        .dq-progress-wrap { background:#ffffff; border:1.5px solid #d9cef0;
            border-radius:12px; padding:1.25rem 1.5rem; margin:1rem 0; }
        .dq-prog-step { display:flex; align-items:center; gap:0.75rem;
            padding:0.45rem 0; border-bottom:1px solid #f0edf8; }
        .dq-prog-step:last-child { border-bottom:none; }
        .dq-prog-dot { width:10px; height:10px; border-radius:50%;
            background:#e5e7eb; flex-shrink:0; }
        .dq-prog-dot.done { background:#10b981; }
        .dq-prog-dot.active { background:#5b2d90;
            box-shadow:0 0 0 3px rgba(91,45,144,0.2); animation:beacon-pulse 1.2s infinite; }
        .dq-prog-label { font-size:0.82rem; color:#4a4a6a; font-weight:500; }
        .dq-prog-label.active { color:#5b2d90; font-weight:700; }
        .dq-prog-label.done { color:#059669; }
        </style>
        """, unsafe_allow_html=True)

        steps_ph = st.empty()

        def _show_steps(active_idx):
            step_defs = [
                ("📂", "Saving & loading files"),
                ("🔧", "Building rulebook"),
                ("⚙️", "Executing validation rules"),
                ("📊", "Calculating DQ scores"),
                ("💾", "Generating Excel report"),
                ("📄", "Building PDF report"),
            ]
            html = '<div class="dq-progress-wrap">'
            for i, (ico, lbl) in enumerate(step_defs):
                if i < active_idx:
                    dot_cls = "done"; lbl_cls = "done"; status = "✓"
                elif i == active_idx:
                    dot_cls = "active"; lbl_cls = "active"; status = ico
                else:
                    dot_cls = ""; lbl_cls = ""; status = ico
                html += f'<div class="dq-prog-step"><div class="dq-prog-dot {dot_cls}"></div><span class="dq-prog-label {lbl_cls}">{status} {lbl}</span></div>'
            html += "</div>"
            steps_ph.markdown(html, unsafe_allow_html=True)

        _show_steps(0)
        data_path  = save_uploaded_file(data_file,  AppConfig.TEMP_DIR)
        rules_path = save_uploaded_file(rules_file, AppConfig.TEMP_DIR)

        _show_steps(0)
        loader = FileLoaderService()
        df     = loader.load_dataframe(data_path, sheet_name=sheet_name)
        cols   = list(df.columns)

        _show_steps(1)
        rb_svc = RulebookBuilderService()
        if rules_file.name.lower().endswith(".json"):
            rulebook = rb_svc.load_json_rulebook(rules_path)
        else:
            rulebook = rb_svc.build_from_rules_dataset(
                loader.load_dataframe(rules_path), cols)

        _show_steps(2)
        executor = RuleExecutorEngine(df, rulebook)
        results  = executor.execute_all_rules()
        combos   = executor.get_combination_duplicates()

        _show_steps(3)
        overall    = ScoringService.calculate_overall_score(results)
        col_scores = ScoringService.calculate_column_scores(results, cols)
        dim_scores = ScoringService.calculate_dimension_scores(results)

        _show_steps(4)
        excel_filename = get_timestamp_filename(f"DQ_Report_{obj_name or 'Dataset'}", "xlsx")
        xl_path        = AppConfig.OUTPUT_DIR / excel_filename
        rgen           = ExcelReportGenerator(
            results_df=results, rulebook=rulebook, all_columns=cols,
            column_scores=col_scores, overall_score=overall,
            dimension_scores=dim_scores, duplicate_combinations=combos,
        )
        rgen.generate_report(AppConfig.OUTPUT_DIR)
        rb_path = rgen.save_rulebook_json(AppConfig.OUTPUT_DIR, rulebook)
        default_rp = AppConfig.OUTPUT_DIR / "DQ_Assessment_Report.xlsx"
        if default_rp.exists() and not xl_path.exists():
            default_rp.rename(xl_path)

        _show_steps(5)
        pdf_bytes = _build_dq_pdf_report(overall, dim_scores, results, col_scores,
                                          obj_name or "Dataset")

        # All done — clear progress, show success banner
        steps_ph.markdown("""
        <div style="background:#f0fdf4;border:1.5px solid #bbf7d0;border-radius:12px;
             padding:0.9rem 1.25rem;display:flex;align-items:center;gap:0.75rem;margin:1rem 0;">
            <span style="font-size:1.2rem;">✅</span>
            <span style="color:#15803d;font-weight:700;font-size:0.9rem;">
                Assessment completed successfully! Dashboard is ready below.
            </span>
        </div>
        """, unsafe_allow_html=True)

        # ── Persist ALL results to session state ─────────────────────────
        st.session_state["dq_score"]          = overall
        st.session_state["dq_dim_scores"]     = dim_scores
        st.session_state["dq_results_df"]     = results
        st.session_state["dq_col_scores"]     = col_scores
        st.session_state["dq_object_name"]    = obj_name or "Customer"
        st.session_state["dq_excel_path"]     = xl_path
        st.session_state["dq_excel_filename"] = excel_filename
        st.session_state["dq_rb_path"]        = str(rb_path) if rb_path else None
        st.session_state["dq_pdf_bytes"]      = pdf_bytes

        st.session_state["mat_objects"] = [obj_name] if obj_name else DEFAULT_MASTER_OBJECTS[:]
        autofill_dq_dimension(overall)

        # ── Render results (also persisted for next visit) ───────────────
        _render_dq_results(
            overall, dim_scores, results, col_scores,
            obj_name or "Dataset", pdf_bytes,
            xl_path, excel_filename,
            str(rb_path) if rb_path else None,
        )

    except Exception as e:
        st.markdown(f"""
        <div class="dash-panel" style="border-color:#fecaca;background:#fef2f2;">
            <div style="color:#dc2626;font-weight:700;">❌ Assessment Error</div>
            <div style="color:#7f1d1d;font-size:0.85rem;margin-top:0.5rem;">{e}</div>
        </div>
        """, unsafe_allow_html=True)
        with st.expander("🔍 Technical Details"):
            st.code(traceback.format_exc())


# ══════════════════════════════════════════════════════════════════════════
#  PAGE: MATURITY ASSESSMENT
# ══════════════════════════════════════════════════════════════════════════
def _apply_editor_edits(dim: str, editor_key: str) -> None:
    widget_state = st.session_state.get(editor_key)
    if not widget_state:
        return
    edited_rows = widget_state.get("edited_rows", {})
    if not edited_rows:
        return
    df = st.session_state.mat_responses[dim].copy()
    for row_idx, changes in edited_rows.items():
        for col, val in changes.items():
            df.at[int(row_idx), col] = val
    st.session_state.mat_responses[dim] = df


def _do_submit() -> None:
    objects   = st.session_state.mat_objects
    dims      = st.session_state.mat_dims
    responses = st.session_state.mat_responses
    cn        = st.session_state.mat_client_name or "Client"
    bm        = float(st.session_state.mat_benchmark)
    tg        = float(st.session_state.mat_target)
    lt        = float(st.session_state.mat_low_thr)
    dq_score  = st.session_state.get("dq_score")

    ok, msg = validate_responses(responses, dims, objects)
    if not ok:
        st.error(f"⚠️ Validation failed: {msg}")
        return

    with st.spinner("⚙️ Computing scores and building reports…"):
        dim_table, overall = compute_all_scores(objects, dims, responses)
        domain_display = {
            dim: float(np.nanmean(dim_table.loc[dim].values)) for dim in dims
        }
        exec_score = float(np.nanmean(overall.values)) if len(overall) else 0.0
        slide_png  = render_slide_png(
            client_name=cn, domain_scores=domain_display,
            exec_score=exec_score if np.isfinite(exec_score) else 0.0,
            benchmark=bm, target=tg,
        )
        pdf_bytes = build_pdf_bytes(
            client_name=cn, slide_png=slide_png, dim_table=dim_table,
            overall=overall, detail_tables=responses, dq_score=dq_score,
        )
        mat_excel = to_excel_bytes(
            dim_table=dim_table, overall=overall, detail_tables=responses,
            low_thr=lt, objects=objects,
        )
        combined_excel = (
            _combined_excel(dq_score, st.session_state.get("dq_dim_scores"), mat_excel)
            if dq_score is not None else mat_excel
        )

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    st.session_state["mat_submitted"] = True
    st.session_state["mat_payload"]   = {
        "dim_table": dim_table, "overall": overall,
        "slide_png": slide_png, "mat_excel": mat_excel,
        "combined_excel": combined_excel, "pdf_bytes": pdf_bytes,
        "client_name": cn, "ts": ts,
    }
    st.rerun()


def page_maturity():
    inject_gdg_light()
    dq_score  = st.session_state.get("dq_score")
    submitted = st.session_state.get("mat_submitted", False)

    with st.sidebar:
        st.markdown("### 🧭 Navigation")
        if st.button("🏠 Home",     use_container_width=True, key="mat_home"):
            st.session_state["page"] = "home"; st.rerun()
        if st.button("🔍 DQ",       use_container_width=True, key="mat_dq"):
            st.session_state["page"] = "dq"; st.rerun()
        if st.button("📋 Policies", use_container_width=True, key="mat_policy"):
            st.session_state["page"] = "policy"; st.rerun()
        if st.button("🎯 Cases",    use_container_width=True, key="mat_cases"):
            st.session_state["page"] = "cases"; st.rerun()
        st.divider()

        st.markdown("### ⚙️ Configuration")
        st.session_state["mat_client_name"] = st.text_input(
            "Client Name", value=st.session_state.get("mat_client_name", ""),
            placeholder="Organisation name", disabled=submitted,
        )
        all_obj_opts = list(dict.fromkeys(
            DEFAULT_MASTER_OBJECTS + st.session_state.mat_objects
        ))
        st.session_state["mat_objects"] = st.multiselect(
            "Master Data Objects", options=all_obj_opts,
            default=st.session_state.mat_objects, disabled=submitted,
        )
        st.session_state["mat_dims"] = st.multiselect(
            "Maturity Dimensions", options=MATURITY_DIMS,
            default=st.session_state.mat_dims, disabled=submitted,
        )
        st.divider()
        st.markdown("### 📊 Thresholds")
        st.session_state["mat_low_thr"] = st.slider(
            "Exception threshold (≤)", 1.0, 5.0,
            float(st.session_state.get("mat_low_thr", 2.0)), 0.5, disabled=submitted,
        )
        st.divider()
        st.markdown("### 🎯 Benchmark / Target")
        st.session_state["mat_benchmark"] = st.number_input(
            "Industry Benchmark", 1.0, 5.0,
            float(st.session_state.get("mat_benchmark", 3.0)), 0.1, disabled=submitted,
        )
        st.session_state["mat_target"] = st.number_input(
            "Target Score", 1.0, 5.0,
            float(st.session_state.get("mat_target", 3.0)), 0.1, disabled=submitted,
        )

    if not st.session_state.mat_objects or not st.session_state.mat_dims:
        st.markdown('<div class="info-box">', unsafe_allow_html=True)
        st.markdown("👉 Please select at least one **Object** and one **Dimension** in the sidebar.")
        st.markdown('</div>', unsafe_allow_html=True)
        st.stop()

    # ── DEFERRED SYNC FIX ──────────────────────────────────────────────────
    # When the user changes objects/dims in the multiselect, Streamlit reruns
    # immediately. If we call sync_response_tables() in that same rerun, it
    # rebuilds DataFrames and pops editor snapshot keys while Streamlit is
    # still reconciling widget state — causing the dropdown to lose focus and
    # discard the selection (requiring multiple clicks).
    #
    # Fix: on the first rerun after a change, just set a _sync_pending flag
    # and trigger another rerun. Only on the second rerun (when widget state
    # is fully settled) do we actually perform the sync.
    # ───────────────────────────────────────────────────────────────────────
    prev_objs = st.session_state.get("_last_sync_objects")
    prev_dims = st.session_state.get("_last_sync_dims")
    curr_objs = st.session_state.mat_objects
    curr_dims = st.session_state.mat_dims

    needs_sync = (
        prev_objs is None
        or prev_dims is None
        or set(prev_objs) != set(curr_objs)
        or set(prev_dims) != set(curr_dims)
    )

    if needs_sync:
        if st.session_state.get("_sync_pending"):
            # Second rerun: widget state is settled, safe to sync now
            sync_response_tables()
            for d in curr_dims:
                st.session_state.pop(f"mat_snap_{d}", None)
            st.session_state["_last_sync_objects"] = list(curr_objs)
            st.session_state["_last_sync_dims"]    = list(curr_dims)
            st.session_state["_sync_pending"] = False
        else:
            # First rerun after change: just flag and rerun again
            st.session_state["_sync_pending"] = True
            st.rerun()

    if dq_score is not None and not st.session_state.get("dq_autofilled"):
        autofill_dq_dimension(dq_score)
        st.session_state.pop("mat_snap_Data Quality", None)
        st.session_state["dq_autofilled"] = True

    # ── REPORT VIEW ────────────────────────────────────────────────────────
    if submitted and st.session_state.get("mat_payload"):
        p  = st.session_state["mat_payload"]
        cn = p["client_name"]
        ts = p["ts"]

        render_uniqus_topbar("Data Maturity Assessment — Report")
        st.markdown("# ✅ Data Maturity Assessment Report")

        if dq_score is not None:
            lvl = dq_score_to_maturity_level(dq_score)
            st.markdown(
                f'<div class="banner success">'
                f'**DQ Engine Score:** {dq_score:.1f}% → **Level:** {lvl} (applied to Data Quality dimension)'
                f'</div>',
                unsafe_allow_html=True,
            )

        st.markdown("### 📊 Summary Slide")
        st.image(p["slide_png"], use_container_width=True)

        UIComponents.render_micro_progress(100, "#10b981", "#34d399")
        st.divider()

        st.markdown("""
            <style>
            .dataframe tbody tr:hover { background-color: rgba(224,242,254,0.5) !important; }
            .dataframe thead th {
                background: linear-gradient(135deg,#e0f2fe 0%,#bae6fd 100%) !important;
                color: #0c4a6e !important;
            }
            </style>
        """, unsafe_allow_html=True)

        t1, t2 = st.columns(2)
        with t1:
            st.markdown("#### Dimension-wise Maturity")
            styled_dim = p["dim_table"].style\
                .format("{:.2f}")\
                .background_gradient(cmap="Blues", axis=None, vmin=1, vmax=5)
            st.dataframe(styled_dim, use_container_width=True)
        with t2:
            st.markdown("#### Overall Maturity Score")
            styled_overall = pd.DataFrame(p["overall"]).T.style\
                .format("{:.2f}")\
                .background_gradient(cmap="Blues", axis=None, vmin=1, vmax=5)
            st.dataframe(styled_overall, use_container_width=True)

        st.divider()
        st.markdown("#### Scores by Dimension")
        dim_vals = {
            dim: float(np.nanmean(p["dim_table"].loc[dim].values))
            for dim in p["dim_table"].index
        }
        bar_img = _mat_bar_png(dim_vals)
        if bar_img:
            st.image(bar_img, use_container_width=True)

        st.divider()
        st.markdown("### 📥 Download Reports")
        safe_cn = cn.replace(" ", "_")

        UIComponents.render_hint_chip(
            "3 formats available", tip="PDF · Maturity Excel · Combined DQ+Maturity Excel", icon="📥"
        )
        st.markdown('<div style="height:0.5rem;"></div>', unsafe_allow_html=True)

        d1, d2, d3 = st.columns(3)
        with d1:
            pdf_filename = get_timestamp_filename(f"Maturity_Report_{safe_cn}", "pdf")
            st.download_button(
                "📄 PDF Report", data=p["pdf_bytes"],
                file_name=pdf_filename, mime="application/pdf",
                use_container_width=True,
            )
        with d2:
            mat_excel_filename = get_timestamp_filename(f"Maturity_Assessment_{safe_cn}", "xlsx")
            st.download_button(
                "📊 Maturity Excel", data=p["mat_excel"],
                file_name=mat_excel_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with d3:
            combined_filename = get_timestamp_filename(f"DQ_Maturity_Combined_{safe_cn}", "xlsx")
            st.download_button(
                "🔗 Combined Excel", data=p["combined_excel"],
                file_name=combined_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        st.divider()
        if st.button("✏️ Edit Responses", use_container_width=True, key="mat_edit"):
            st.session_state["mat_submitted"] = False
            st.session_state["mat_payload"]   = {}
            st.rerun()
        st.stop()

    # ── QUESTIONNAIRE VIEW ─────────────────────────────────────────────────
    render_uniqus_topbar("Data Maturity Assessment")
    st.markdown("# 📈 Data Maturity Assessment")

    if dq_score is not None:
        lvl = dq_score_to_maturity_level(dq_score)
        st.markdown(
            f'<div class="banner success">'
            f'✅ DQ Score **{dq_score:.1f}%** → level **{lvl}** auto-filled in *Data Quality* dimension.'
            f'</div>',
            unsafe_allow_html=True,
        )

    UIComponents.render_action_hint_bar(
        title="How to complete",
        message="Select a rating for each question using the "
                "<strong>dropdown columns</strong>. Weighted scoring is automatic.",
        color="#a78bfa",
    )
    st.divider()

    dims = st.session_state.mat_dims
    tabs = st.tabs(dims)

    for i, dim in enumerate(dims):
        with tabs[i]:
            st.markdown(f"### {dim}")
            if dim == "Data Quality" and dq_score is not None:
                lvl = dq_score_to_maturity_level(dq_score)
                st.markdown(
                    f'<div class="banner">'
                    f'Auto-populated from DQ Score **{dq_score:.1f}%** → **{lvl}**. '
                    f'You can adjust individual ratings as needed.'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            cfg = {"Weight": st.column_config.NumberColumn("Weight", min_value=0.0, step=0.5)}
            for obj in st.session_state.mat_objects:
                cfg[obj] = st.column_config.SelectboxColumn(obj, options=RATING_LABELS, required=True)

            editor_key = f"mat_editor_{dim}"
            st.data_editor(
                st.session_state.mat_responses[dim],
                use_container_width=True, hide_index=True,
                column_config=cfg,
                disabled=["Question ID", "Section", "Question"],
                key=editor_key,
                on_change=_apply_editor_edits,
                args=(dim, editor_key),
            )

    st.divider()
    col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
    with col2:
        if st.button("🚀 Submit & Generate Report", type="primary",
                     use_container_width=True, key="mat_submit"):
            _do_submit()
    with col3:
        st.info("**Submit** to generate visuals and downloadable reports.")


# ══════════════════════════════════════════════════════════════════════════
#  PAGE: POLICY HUB
# ══════════════════════════════════════════════════════════════════════════
def page_policy_hub():
    with st.sidebar:
        st.markdown("### 🧭 Navigation")
        if st.button("🏠 Home",      use_container_width=True, key="policy_home"):
            st.session_state["page"] = "home"; st.rerun()
        if st.button("🔍 DQ",        use_container_width=True, key="policy_dq"):
            st.session_state["page"] = "dq"; st.rerun()
        if st.button("📈 Maturity",  use_container_width=True, key="policy_maturity"):
            st.session_state["page"] = "maturity"; st.rerun()
        if st.button("🎯 Cases",     use_container_width=True, key="policy_cases"):
            st.session_state["page"] = "cases"; st.rerun()

    # ── Top bar ───────────────────────────────────────────────────────────
    render_uniqus_topbar("Policy Hub & Procedures Management")

    # ── Hero ─────────────────────────────────────────────────────────────
    st.markdown("""
    <div class="uniqus-hero" style="padding:2.2rem 2.5rem 2rem;">
        <div class="uniqus-hero-badge">📋 Enterprise Governance</div>
        <h1 style="font-size:2rem!important;">Policy Hub & Procedures Management</h1>
        <p style="margin-bottom:0;">Centralized repository for enterprise data governance policies,
        procedures, approvals and compliance tracking — all in one place.</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Platform KPI strip ────────────────────────────────────────────────
    st.markdown("""
    <div class="quick-stat-bar" style="margin-bottom:2rem;">
        <div class="quick-stat-item">
            <div class="quick-stat-val">4</div>
            <div class="quick-stat-lbl">Core Modules</div>
        </div>
        <div class="quick-stat-item">
            <div class="quick-stat-val magenta">4</div>
            <div class="quick-stat-lbl">Lifecycle Stages</div>
        </div>
        <div class="quick-stat-item">
            <div class="quick-stat-val teal">4</div>
            <div class="quick-stat-lbl">User Roles</div>
        </div>
        <div class="quick-stat-item">
            <div class="quick-stat-val" style="color:#d97706;">SSO</div>
            <div class="quick-stat-lbl">Azure AD Login</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════
    # SECTION 1 — Policy Lifecycle Pipeline
    # ══════════════════════════════════════════════════════════════════════
    st.markdown("""
    <div class="dash-section-header">
        <div class="dash-section-dot"></div>
        <h3>Policy Lifecycle Pipeline</h3>
        <div class="dash-section-accent"></div>
    </div>
    <div class="pol-lifecycle">
        <div class="pol-stage" data-stage="1">
            <div class="pol-stage-num">01</div>
            <div class="pol-stage-icon">✏️</div>
            <div class="pol-stage-label">Draft</div>
            <div class="pol-stage-desc">Policy authored &amp; saved as draft by document owner</div>
        </div>
        <div class="pol-stage-arrow">→</div>
        <div class="pol-stage" data-stage="2">
            <div class="pol-stage-num">02</div>
            <div class="pol-stage-icon">👁️</div>
            <div class="pol-stage-label">Under Review</div>
            <div class="pol-stage-desc">Sent to reviewers via one-click submit; email links generated</div>
        </div>
        <div class="pol-stage-arrow">→</div>
        <div class="pol-stage active" data-stage="3">
            <div class="pol-stage-num">03</div>
            <div class="pol-stage-icon">✅</div>
            <div class="pol-stage-label">Approved</div>
            <div class="pol-stage-desc">All approvers have signed off; escalation alerts handled</div>
        </div>
        <div class="pol-stage-arrow">→</div>
        <div class="pol-stage" data-stage="4">
            <div class="pol-stage-num">04</div>
            <div class="pol-stage-icon">📢</div>
            <div class="pol-stage-label">Published</div>
            <div class="pol-stage-desc">Live in repository; stakeholders notified automatically</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════
    # SECTION 2 — Four Core Module Cards (2×2 grid)
    # ══════════════════════════════════════════════════════════════════════
    st.markdown("""
    <div class="dash-section-header" style="margin-top:2rem;">
        <div class="dash-section-dot magenta"></div>
        <h3>Platform Capability Modules</h3>
        <div class="dash-section-accent"></div>
    </div>
    """, unsafe_allow_html=True)

    c1, c2 = st.columns(2, gap="large")

    with c1:
        # Module 1 — Workflow Automation
        st.markdown("""
        <div class="pol-module-card purple">
            <div class="pol-mod-header">
                <div class="pol-mod-icon-box purple">⚙️</div>
                <div>
                    <div class="pol-mod-title">Workflow Automation</div>
                    <div class="pol-mod-subtitle">End-to-end approval orchestration</div>
                </div>
                <span class="pol-mod-badge live">Live</span>
            </div>
            <div class="pol-cap-grid">
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">🚀</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">One-Click Submit</div>
                        <div class="pol-cap-text">Send policies to reviewers instantly — no manual routing required.</div>
                    </div>
                </div>
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">📊</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">Status Tracker</div>
                        <div class="pol-cap-text">Visual pipeline showing current stage across the full lifecycle.</div>
                    </div>
                </div>
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">🕐</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">Approval Timeline</div>
                        <div class="pol-cap-text">Full audit log — who reviewed, approved or rejected and when.</div>
                    </div>
                </div>
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">📧</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">Email Approval Links</div>
                        <div class="pol-cap-text">Approve or reject directly from inbox — no portal login needed.</div>
                    </div>
                </div>
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">⚠️</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">Escalation Alerts</div>
                        <div class="pol-cap-text">Delayed approvals are auto-flagged and escalated up the chain.</div>
                    </div>
                </div>
            </div>
            <div class="pol-mod-benefit">
                <span class="pol-benefit-dot purple"></span>
                <span><strong>Outcome:</strong> Zero manual tracking — fully automated, always visible.</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

    with c2:
        # Module 2 — Notifications & Reminders
        st.markdown("""
        <div class="pol-module-card magenta">
            <div class="pol-mod-header">
                <div class="pol-mod-icon-box magenta">🔔</div>
                <div>
                    <div class="pol-mod-title">Notifications &amp; Reminders</div>
                    <div class="pol-mod-subtitle">Proactive stakeholder communication</div>
                </div>
                <span class="pol-mod-badge live">Live</span>
            </div>
            <div class="pol-cap-grid">
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">🔔</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">In-App Bell</div>
                        <div class="pol-cap-text">Real-time alerts inside the portal with unread count badge.</div>
                    </div>
                </div>
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">📬</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">Notification Feed</div>
                        <div class="pol-cap-text">Policy approved · Review requested · Comments added — live stream.</div>
                    </div>
                </div>
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">📧</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">Email Alerts</div>
                        <div class="pol-cap-text">All policy events pushed directly to Outlook / corporate mail.</div>
                    </div>
                </div>
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">⏰</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">Smart Reminders</div>
                        <div class="pol-cap-text">Proactive nudges for pending approvals, overdue tasks, review dates.</div>
                    </div>
                </div>
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">⚙️</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">Digest Settings</div>
                        <div class="pol-cap-text">Users choose: instant · daily digest · weekly summary.</div>
                    </div>
                </div>
            </div>
            <div class="pol-mod-benefit">
                <span class="pol-benefit-dot magenta"></span>
                <span><strong>Outcome:</strong> No missed approvals or deadlines — ever.</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    c3, c4 = st.columns(2, gap="large")

    with c3:
        # Module 3 — Role-Based Access
        st.markdown("""
        <div class="pol-module-card teal">
            <div class="pol-mod-header">
                <div class="pol-mod-icon-box teal">🔐</div>
                <div>
                    <div class="pol-mod-title">Role-Based User Access</div>
                    <div class="pol-mod-subtitle">Governed permissions &amp; security</div>
                </div>
                <span class="pol-mod-badge live">Live</span>
            </div>
            <div class="pol-roles-row">
                <div class="pol-role-chip admin">👑 Admin</div>
                <div class="pol-role-chip editor">✏️ Editor</div>
                <div class="pol-role-chip reviewer">👁️ Reviewer</div>
                <div class="pol-role-chip viewer">📖 Viewer</div>
            </div>
            <div class="pol-cap-grid" style="margin-top:0.75rem;">
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">🏠</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">Role-Based Dashboard</div>
                        <div class="pol-cap-text">Personalised homepage content per user role automatically.</div>
                    </div>
                </div>
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">🔒</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">Restricted Document View</div>
                        <div class="pol-cap-text">Sensitive policies visible only to authorised clearance levels.</div>
                    </div>
                </div>
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">🏢</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">Department Filtering</div>
                        <div class="pol-cap-text">Users automatically see only their department's relevant policies.</div>
                    </div>
                </div>
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">🔑</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">SSO via Azure AD</div>
                        <div class="pol-cap-text">Seamless login with existing company credentials — no new passwords.</div>
                    </div>
                </div>
            </div>
            <div class="pol-mod-benefit">
                <span class="pol-benefit-dot teal"></span>
                <span><strong>Outcome:</strong> Maximum security with minimum friction for users.</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

    with c4:
        # Module 4 — White-Labelling
        st.markdown("""
        <div class="pol-module-card amber">
            <div class="pol-mod-header">
                <div class="pol-mod-icon-box amber">🎨</div>
                <div>
                    <div class="pol-mod-title">White-Label Branding</div>
                    <div class="pol-mod-subtitle">Full corporate identity alignment</div>
                </div>
                <span class="pol-mod-badge live">Live</span>
            </div>
            <div class="pol-cap-grid">
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">🏷️</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">Logo &amp; Brand Colours</div>
                        <div class="pol-cap-text">Portal displays company logo, corporate palette and approved typography.</div>
                    </div>
                </div>
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">🏠</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">Custom Homepage</div>
                        <div class="pol-cap-text">Dashboard layout configured to specific business structure and needs.</div>
                    </div>
                </div>
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">📧</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">Branded Email Templates</div>
                        <div class="pol-cap-text">All outgoing notifications follow company branding guidelines.</div>
                    </div>
                </div>
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">🌗</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">Light / Dark Theme</div>
                        <div class="pol-cap-text">User-selectable theme for comfortable, accessible viewing.</div>
                    </div>
                </div>
                <div class="pol-cap-item">
                    <div class="pol-cap-icon">🧩</div>
                    <div class="pol-cap-body">
                        <div class="pol-cap-name">Personalised Widgets</div>
                        <div class="pol-cap-text">My Tasks · Recent Policies · Pending Approvals — user-configurable.</div>
                    </div>
                </div>
            </div>
            <div class="pol-mod-benefit">
                <span class="pol-benefit-dot amber"></span>
                <span><strong>Outcome:</strong> The tool feels entirely your own — trusted and familiar.</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════
    # SECTION 3 — Platform Summary Table
    # ══════════════════════════════════════════════════════════════════════
    st.markdown("""
    <div class="dash-section-header" style="margin-top:2rem;">
        <div class="dash-section-dot"></div>
        <h3>Capability Summary</h3>
        <div class="dash-section-accent"></div>
    </div>
    <div class="dash-panel">
        <table class="score-table" style="font-size:0.82rem;">
            <thead>
                <tr>
                    <th style="width:22%;">Module</th>
                    <th style="width:30%;">Key Capabilities</th>
                    <th style="width:28%;">User Benefit</th>
                    <th style="width:20%;">Status</th>
                </tr>
            </thead>
            <tbody>
                <tr>
                    <td><strong>⚙️ Workflow Automation</strong></td>
                    <td>Submit, track, approve, escalate</td>
                    <td>Zero manual tracking</td>
                    <td><span class="score-pill good">✅ Live</span></td>
                </tr>
                <tr>
                    <td><strong>🔔 Notifications</strong></td>
                    <td>In-app, email, digest, reminders</td>
                    <td>No missed deadlines</td>
                    <td><span class="score-pill good">✅ Live</span></td>
                </tr>
                <tr>
                    <td><strong>🔐 Role-Based Access</strong></td>
                    <td>Admin · Editor · Reviewer · Viewer</td>
                    <td>Secure &amp; clutter-free UX</td>
                    <td><span class="score-pill good">✅ Live</span></td>
                </tr>
                <tr>
                    <td><strong>🎨 White-Labelling</strong></td>
                    <td>Logo, colours, custom homepage</td>
                    <td>Fully branded experience</td>
                    <td><span class="score-pill good">✅ Live</span></td>
                </tr>
            </tbody>
        </table>
    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
#  START APPLICATION
# ══════════════════════════════════════════════════════════════════════════
load_css()
_init_state()

{
    "home":     page_home,
    "dq":       page_dq,
    "maturity": page_maturity,
    "policy":   page_policy_hub,
    "cases":    page_case_management,
}[st.session_state.page]()
